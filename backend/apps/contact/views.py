from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Q

from .models import Contact
from .serializers import ContactSerializer
from rest_framework import generics
from rest_framework.permissions import IsAdminUser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from django.utils import timezone

class ContactDetailAPIView(generics.RetrieveAPIView):
    """
    Get a single contact enquiry.
    """

    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAdminUser]

class ContactStatusUpdateAPIView(generics.UpdateAPIView):
    """
    Update only the status of a contact enquiry.
    """

    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAdminUser]
    http_method_names = ["patch"]

    def patch(self, request, *args, **kwargs):
        contact = self.get_object()

        status_value = request.data.get("status")

        if status_value not in ["pending", "completed"]:
            return Response(
                {"error": "Invalid status."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact.status = status_value
        contact.save(update_fields=["status"])

        serializer = self.get_serializer(contact)

        return Response(serializer.data)
    
class ContactExportAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):

        queryset = Contact.objects.all().order_by("-created_at")

        search = request.GET.get("search")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )

        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)

        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        wb = Workbook()
        ws = wb.active

        ws.title = "Contact Enquiries"

        headers = [
            "Name",
            "Email",
            "Phone",
            "Requirement",
            "Status",
            "Submitted On",
        ]

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid",
        )

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col)

            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        row = 2

        for enquiry in queryset:

            ws.cell(row=row, column=1).value = enquiry.name
            ws.cell(row=row, column=2).value = enquiry.email
            ws.cell(row=row, column=3).value = enquiry.phone
            ws.cell(row=row, column=4).value = enquiry.requirement
            ws.cell(row=row, column=5).value = enquiry.status
            ws.cell(row=row, column=6).value = enquiry.created_at.strftime("%d-%m-%Y %H:%M")

            row += 1

        for column_cells in ws.columns:

            length = max(len(str(cell.value or "")) for cell in column_cells)

            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        filename = f"Contact_Enquiries_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)

        return response

class ContactAPIView(generics.ListCreateAPIView):
    serializer_class = ContactSerializer

    def get_permissions(self):
        if self.request.method == "GET":
            return [IsAdminUser()]
        return []

    def get_queryset(self):
        queryset = Contact.objects.all().order_by("-created_at")

        search = self.request.query_params.get("search")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(phone__icontains=search)
            )

        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)

        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(
            {
                "message": "Contact enquiry submitted successfully.",
                "data": serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )

class ContactDeleteAPIView(generics.DestroyAPIView):
    """
    Delete a contact enquiry.
    Only staff/admin users can delete enquiries.
    """

    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAdminUser]

# class ContactCreateAPIView(APIView):
#     """
#     API to create a new contact enquiry.
#     """

#     def post(self, request):
#         serializer = ContactSerializer(data=request.data)

#         if serializer.is_valid():
#             serializer.save()

#             return Response(
#                 {
#                     "message": "Contact enquiry submitted successfully.",
#                     "data": serializer.data,
#                 },
#                 status=status.HTTP_201_CREATED,
#             )

#         return Response(
#             serializer.errors,
#             status=status.HTTP_400_BAD_REQUEST,
#         )
